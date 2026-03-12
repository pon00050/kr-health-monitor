"""
NHIS aggregate data — file-based parsers + stub replacements.

IMPORTANT: All NHIS datasets are BULK FILE DOWNLOADS, not REST APIs.
Individual-level NHIS data (NHISS cohort) requires IRB and on-site access — out of scope.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd

from src.config import NHIS_REGION_MAP, REGION_CODE_REVERSE, REGION_CODES, SIDO_NAME_MAP

logger = logging.getLogger(__name__)

# Aliases for region code bridging: NHIS codes → names → HIRA codes
_NHIS_CODE_TO_NAME = NHIS_REGION_MAP     # "41" → "경기"
_NAME_TO_HIRA_CODE = REGION_CODE_REVERSE  # "경기" → "31"


def parse_yearbook_ch06(paths: list[str | Path]) -> pd.DataFrame:
    """Parse 건강보험통계연보 ch06 disease-classification Excel files.

    Scans the 다빈도 상병 진료현황 (frequent disease) section (sheets 6-3 family)
    for E10 (T1D), E11 (T2D), and E14 (unspecified diabetes) totals.

    Each row in the 6-3 sheets contains TWO disease entries side by side:
      Cols 0-7:  rank, ICD, name, patients, visits, reimb_days, cost_krw_thousands, benefit
      Cols 8-15: same for second entry

    Units: 진료비/급여비 are already in 천원 (1,000 KRW).

    Args:
        paths: List of paths to ch06 Excel files (one per year).
               Year is extracted from the parent directory name.

    Returns DataFrame with columns:
        year, icd_code, patient_count, visit_days, cost_krw_thousands,
        case_count, source
    """
    TARGET_ICD = {"E10", "E11", "E14"}
    # Sheets in the 6-3 family (total inpatient+outpatient) — scan all
    _6_3_PATTERNS = ("6-3", "6-3(")

    all_rows: list[dict] = []

    for raw_path in paths:
        path = Path(raw_path)
        if not path.exists():
            logger.warning(f"ch06 file not found: {path}")
            continue

        # Extract year from parent directory name (e.g. "2022_건강보험통계연보_본문" → 2022)
        year = _extract_year_from_path(path)
        if year is None:
            logger.warning(f"Cannot determine year from path: {path}")
            continue

        logger.info(f"Parsing ch06 yearbook for year {year}: {path.name}")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
        except Exception as e:
            logger.error(f"Failed to open {path}: {e}")
            continue

        found: dict[str, dict] = {}  # icd_code → row dict

        for sheet_name in wb.sheetnames:
            # Only scan 6-3 family (total figures, not inpatient/outpatient splits)
            if not (sheet_name == "6-3" or sheet_name.startswith("6-3(")):
                continue
            if len(found) == len(TARGET_ICD):
                break

            ws = wb[sheet_name]
            for row_vals in ws.iter_rows(values_only=True):
                if len(found) == len(TARGET_ICD):
                    break
                row = list(row_vals)
                # Each row has two entries (first: cols 0-7, second: cols 8-15)
                for offset in (0, 8):
                    if offset + 7 >= len(row):
                        continue
                    icd = str(row[offset + 1]).strip() if row[offset + 1] is not None else ""
                    if icd in TARGET_ICD and icd not in found:
                        try:
                            found[icd] = {
                                "year": year,
                                "icd_code": icd,
                                "patient_count": int(row[offset + 3]),
                                "visit_days": int(row[offset + 4]),
                                "case_count": int(row[offset + 5]),   # 요양급여일수
                                "cost_krw_thousands": float(row[offset + 6]),
                                "source": "nhis_yearbook_ch06",
                            }
                        except (TypeError, ValueError):
                            pass

        wb.close()
        all_rows.extend(found.values())

    if not all_rows:
        return pd.DataFrame(columns=["year", "icd_code", "patient_count", "visit_days",
                                      "cost_krw_thousands", "case_count", "source"])
    return pd.DataFrame(all_rows)


def parse_regional_utilization_excel(paths: list[str | Path]) -> pd.DataFrame:
    """Parse 지역별의료이용통계연보 05_질병진료 Excel files for 당뇨(E10-E14) data.

    Scans ALL sheets in each file for 시도-level diabetes benefit rows.
    시도 summary rows are identified by exact name match in REGION_CODES values.

    File structure (sheet 690p and others):
      Col 0: 구분 (region name — 시도 summary or 시군구 district)
      Cols 1-5: 고혈압 data
      Col 6: 당뇨 진료실인원수 (patients)
      Col 7: 당뇨 입내원일수 (visits)
      Col 8: 당뇨 요양급여일수 (reimbursed days)
      Col 9: 당뇨 진료비 (medical expense, 천원)
      Col 10: 당뇨 급여비 (benefit, 천원)

    Args:
        paths: List of paths to 05_질병진료 Excel files (one per year).

    Returns DataFrame with columns:
        region_code, region_name, year, patient_count, cost_krw_thousands, source
    """
    all_rows: list[dict] = []

    for raw_path in paths:
        path = Path(raw_path)
        if not path.exists():
            logger.warning(f"Regional utilization file not found: {path}")
            continue

        year = _extract_year_from_path(path)
        if year is None:
            logger.warning(f"Cannot determine year from path: {path}")
            continue

        logger.info(f"Parsing regional utilization for year {year}: {path.name}")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
        except Exception as e:
            logger.error(f"Failed to open {path}: {e}")
            continue

        found: set[str] = set()

        for sheet_name in wb.sheetnames:
            if len(found) >= len(REGION_CODES):
                break  # All 17 regions found — no need to scan more sheets
            ws = wb[sheet_name]
            for row_vals in ws.iter_rows(values_only=True):
                row = list(row_vals)
                if not row or row[0] is None:
                    continue
                region_name = str(row[0]).strip()
                if region_name not in REGION_CODE_REVERSE:
                    continue
                region_code = REGION_CODE_REVERSE[region_name]
                if region_code in found:
                    continue  # Already got this region from this file
                try:
                    patient_count = int(row[6])
                    cost_krw_thousands = float(row[9])
                    found.add(region_code)
                    all_rows.append({
                        "region_code": region_code,
                        "region_name": region_name,
                        "year": year,
                        "patient_count": patient_count,
                        "cost_krw_thousands": cost_krw_thousands,
                        "source": "nhis_regional_utilization",
                    })
                except (TypeError, ValueError, IndexError):
                    continue

        wb.close()

    if not all_rows:
        return pd.DataFrame(columns=["region_code", "region_name", "year",
                                      "patient_count", "cost_krw_thousands", "source"])
    return pd.DataFrame(all_rows)


def parse_checkup_csv(path: str | Path) -> pd.DataFrame:
    """Parse 국민건강보험공단_건강검진정보 CSV for blood glucose by region.

    Actual file:
      Encoding: cp949, ~1M rows
      Key columns: 시도코드 (NHIS internal codes), 식전혈당(공복혈당)
      Year: from filename (e.g. 건강검진정보_2024.CSV → 2024)

    Applies NHIS_REGION_MAP to translate NHIS codes to standard region names,
    then maps to HIRA 2-digit region codes via REGION_CODES.

    Returns 17-row DataFrame with columns:
        region_code, region_name, year, mean_fasting_glucose,
        high_glucose_rate_pct, screened_count, source
    """
    path = Path(path)
    year = _extract_year_from_path(path)
    if year is None:
        year = 2024  # Fallback

    try:
        df = pd.read_csv(
            path,
            encoding="cp949",
            usecols=["시도코드", "식전혈당(공복혈당)"],
            dtype={"시도코드": str},
        )
    except Exception as e:
        logger.error(f"Failed to read checkup CSV: {e}")
        return pd.DataFrame(columns=["region_code", "region_name", "year",
                                      "mean_fasting_glucose", "high_glucose_rate_pct",
                                      "screened_count", "source"])

    df.columns = ["nhis_code", "fasting_glucose"]
    df["fasting_glucose"] = pd.to_numeric(df["fasting_glucose"], errors="coerce")
    df = df.dropna(subset=["fasting_glucose"])
    df["nhis_code"] = df["nhis_code"].str.strip()

    rows = []
    for nhis_code, group in df.groupby("nhis_code"):
        region_name = _NHIS_CODE_TO_NAME.get(str(nhis_code))
        if region_name is None:
            continue
        region_code = _NAME_TO_HIRA_CODE.get(region_name)
        if region_code is None:
            continue
        glucose = group["fasting_glucose"]
        rows.append({
            "region_code": region_code,
            "region_name": region_name,
            "year": year,
            "mean_fasting_glucose": round(glucose.mean(), 2),
            "high_glucose_rate_pct": round((glucose >= 126).sum() / len(glucose) * 100, 2),
            "screened_count": len(glucose),
            "source": f"nhis_checkup_{year}",
        })

    if not rows:
        return pd.DataFrame(columns=["region_code", "region_name", "year",
                                      "mean_fasting_glucose", "high_glucose_rate_pct",
                                      "screened_count", "source"])
    return pd.DataFrame(rows)


def parse_t1d_age_sex_csv(path: str | Path) -> pd.DataFrame:
    """Parse 국민건강보험공단_제1형 당뇨병 환자 수 CSV by year × age × sex.

    Source: NHIS direct file download (20241231 version). Covers 2021–2024.
    Encoding: cp949.

    Suppression note: cells with `*` represent patient counts < 5 (privacy suppression).
    These are stored as NaN — NOT as 0 — to avoid misrepresenting suppressed data.

    Output schema:
        year (int), age (int), sex (str: M/F), patient_count (float, nullable),
        suppressed (bool), source (str)
    """
    path = Path(path)

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read T1D CSV: {e}")
        return pd.DataFrame(columns=["year", "age", "sex", "patient_count", "suppressed", "source"])

    # Normalize column names — strip whitespace
    df.columns = [c.strip() for c in df.columns]

    # Rename columns
    col_map = {
        "진료개시년도": "year",
        "연령": "age_str",
        "성별": "sex_raw",
        "환자수": "patient_count_raw",
    }
    df = df.rename(columns=col_map)

    # Parse year
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")

    # Parse age: strip "살" suffix → int
    df["age"] = (
        df["age_str"].str.replace("살", "", regex=False).str.strip()
    )
    df["age"] = pd.to_numeric(df["age"], errors="coerce").astype("Int64")

    # Map sex
    df["sex"] = df["sex_raw"].map({"남자": "M", "여자": "F"})

    # Parse patient_count: suppress `*` → NaN
    df["suppressed"] = df["patient_count_raw"].str.strip() == "*"
    df["patient_count"] = df["patient_count_raw"].where(
        ~df["suppressed"], other=pd.NA
    )
    df["patient_count"] = pd.to_numeric(df["patient_count"], errors="coerce")

    df["source"] = "NHIS_T1D_AGE_SEX_20241231"

    return df[["year", "age", "sex", "patient_count", "suppressed", "source"]].reset_index(drop=True)


def parse_consumables_monthly_csv(path: str | Path) -> pd.DataFrame:
    """Parse 국민건강보험공단_현금급여비(당뇨병환자소모성재료) 지급현황 CSV.

    Handles all four annual schemas (2021–2024). Column detection uses substring matching
    so the function works regardless of which year's file is passed.

    Schema variants:
      2021: date column '구분' (YYYY-MM), '건수(건)', '금액(원)'
      2022: date column '연월' (YYYY-MM), '건수(건)', '금액(원)'
      2023: '지급년도', '해당월', '지급건수', '지급금액'
      2024: '지급연도', '지급월', '지급건수(건)', '지급금액(원)'

    IMPORTANT INTERPRETIVE CAVEATS:
    1. SCOPE IS BROADER THAN CGM. 소모성재료 급여 covers CGM sensors (T1D, post-2022-08)
       AND test strips/lancets (T1D + insulin-dependent T2D). Cannot be attributed to CGM alone.
    2. transaction_count ≠ unique beneficiaries. T2D consumable claims inflate the count.
    3. Quarterly CGM payments are processed unevenly across months.

    Output schema:
        year (int), month (int), transaction_count (int), payment_won (int), source (str)
    """
    path = Path(path)
    logger.info(f"Parsing consumables monthly CSV: {path.name}")

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read consumables CSV: {e}")
        return pd.DataFrame(columns=["year", "month", "transaction_count", "payment_won", "source"])

    df.columns = [c.strip() for c in df.columns]

    # Find transaction count column (contains '건수')
    txn_col = next((c for c in df.columns if "건수" in c), None)
    # Find payment column (contains '금액')
    pay_col = next((c for c in df.columns if "금액" in c), None)

    if txn_col is None or pay_col is None:
        logger.error(f"Cannot find required columns in {path.name}. Columns: {list(df.columns)}")
        return pd.DataFrame(columns=["year", "month", "transaction_count", "payment_won", "source"])

    df = df.rename(columns={txn_col: "transaction_count", pay_col: "payment_won"})

    # Strip commas from numeric fields
    for col in ["transaction_count", "payment_won"]:
        df[col] = df[col].str.replace(",", "", regex=False).str.strip()

    # Detect date column: a column whose first non-null value matches YYYY-MM pattern
    date_col = None
    for c in df.columns:
        if c in ("transaction_count", "payment_won"):
            continue
        sample = df[c].dropna()
        if not sample.empty and re.match(r"^\d{4}-\d{2}$", str(sample.iloc[0]).strip()):
            date_col = c
            break

    if date_col is not None:
        # 2021/2022 schemas: extract year and month from YYYY-MM values
        date_parsed = df[date_col].str.strip().str.extract(r"^(\d{4})-(\d{2})$")
        df["year"] = pd.to_numeric(date_parsed[0], errors="coerce")
        df["month"] = pd.to_numeric(date_parsed[1], errors="coerce")
    else:
        # 2023/2024 schemas: explicit year and month columns
        year_col = next((c for c in df.columns if "년도" in c or "연도" in c), None)
        # Month column: contains '월' but is not the date column and not '연월'
        month_col = next(
            (c for c in df.columns
             if "월" in c and "년도" not in c and "연도" not in c and "연월" not in c
             and c not in ("transaction_count", "payment_won")),
            None
        )
        if year_col and month_col:
            df["year"] = pd.to_numeric(df[year_col].str.strip(), errors="coerce")
            df["month"] = pd.to_numeric(df[month_col].str.strip(), errors="coerce")
        else:
            logger.error(f"Cannot determine year/month from {path.name}. Columns: {list(df.columns)}")
            df["year"] = pd.NA
            df["month"] = pd.NA

    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    df["month"] = pd.to_numeric(df["month"], errors="coerce").astype("Int64")
    df["transaction_count"] = pd.to_numeric(df["transaction_count"], errors="coerce").astype("Int64")
    df["payment_won"] = pd.to_numeric(df["payment_won"], errors="coerce").astype("Int64")

    df["source"] = "NHIS_CONSUMABLES_MONTHLY"

    # Keep only rows with valid year and month
    df = df[df["year"].notna() & df["month"].notna()].copy()

    return df[["year", "month", "transaction_count", "payment_won", "source"]].reset_index(drop=True)


def parse_cgm_utilization_csv(path: str | Path) -> pd.DataFrame:
    """Parse CGM utilization CSV — unique CGM users by year (2020–2024).

    실수진자 = unique patients by 주민등록번호 (deduplicated). T1D 요양비 claimants only.
    Does NOT include private-pay CGM users (negligible given cost).

    Source: NHIS_CGM_UTILIZATION_20241231. 5 rows (2020–2024).
    Output: year (int), cgm_users (int), source (str)
    """
    path = Path(path)

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read CGM utilization CSV: {e}")
        return pd.DataFrame(columns=["year", "cgm_users", "source"])

    df.columns = [c.strip() for c in df.columns]

    year_col = next((c for c in df.columns if "연도" in c or "년도" in c), None)
    users_col = next((c for c in df.columns if "실수진자" in c or ("수진자" in c and "실" in c)), None)

    if year_col is None or users_col is None:
        logger.error(f"Cannot find required columns in {path.name}. Columns: {list(df.columns)}")
        return pd.DataFrame(columns=["year", "cgm_users", "source"])

    df = df.rename(columns={year_col: "year", users_col: "cgm_users"})

    df["cgm_users"] = df["cgm_users"].str.replace(",", "", regex=False).str.strip()
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    df["cgm_users"] = pd.to_numeric(df["cgm_users"], errors="coerce").astype("Int64")
    df["source"] = "NHIS_CGM_UTILIZATION_20241231"

    df = df[df["year"].notna() & df["cgm_users"].notna()].copy()
    return df[["year", "cgm_users", "source"]].reset_index(drop=True)


def _read_nhis_xlsx_sheet(path: Path, sheet_idx: int) -> pd.DataFrame:
    """Read an NHIS XLSX sheet, auto-detecting the header row.

    NHIS XLSX files typically have 1–3 title rows before the actual column headers.
    Reads with header=None, finds the first row with Korean or English text (not all NaN),
    uses it as the header, and returns data rows below it.

    Returns DataFrame with string columns, or empty DataFrame on failure.
    """
    try:
        df_raw = pd.read_excel(path, sheet_name=sheet_idx, header=None, dtype=str, engine="openpyxl")
    except Exception as e:
        logger.error(f"Failed to read sheet {sheet_idx} from {path.name}: {e}")
        return pd.DataFrame()

    # Find the header row: first row where at least 2 cells are non-null and non-numeric
    header_row_idx = None
    for i, row in df_raw.iterrows():
        non_null = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
        # A header row typically has several string labels; skip rows that are all numbers
        text_cells = [v for v in non_null if not re.match(r"^\d+$", v)]
        if len(text_cells) >= 2:
            header_row_idx = i
            break

    if header_row_idx is None:
        logger.warning(f"Could not find header row in sheet {sheet_idx} of {path.name}")
        return pd.DataFrame()

    # Build column names from the header row
    header_vals = [str(v).strip() if pd.notna(v) else f"col_{j}"
                   for j, v in enumerate(df_raw.iloc[header_row_idx])]

    # Take data rows below the header
    df = df_raw.iloc[header_row_idx + 1:].copy()
    df.columns = header_vals
    df = df.reset_index(drop=True)
    return df


def parse_yoyangbi_registered_xlsx(path: str | Path) -> pd.DataFrame:
    """Parse Sheet 2 of annual diabetes info XLSX for registered 요양비 beneficiary pools.

    Sheet: '연도별 요양비(1형 및 2형 당뇨병) 등록환자 수'
    Source: NHIS_YOYANGBI_REGISTERED_20241231. 6 rows (2019–2024).
    Output: year (int), t1d_registered (int), t2d_registered (int), source (str)
    """
    path = Path(path)

    df_raw = _read_nhis_xlsx_sheet(path, sheet_idx=1)
    if df_raw.empty:
        return pd.DataFrame(columns=["year", "t1d_registered", "t2d_registered", "source"])

    year_col = next((c for c in df_raw.columns if "년도" in c or "연도" in c), None)
    type_col = next((c for c in df_raw.columns if "주상병" in c or "병명" in c or "구분" in c), None)
    count_col = next((c for c in df_raw.columns if "등록환자" in c or ("등록" in c and "환자" in c)), None)

    if not all([year_col, type_col, count_col]):
        logger.warning(f"Column detection failed. Columns: {list(df_raw.columns)} — trying positional")
        if len(df_raw.columns) >= 3:
            year_col, type_col, count_col = df_raw.columns[0], df_raw.columns[1], df_raw.columns[2]
        else:
            return pd.DataFrame(columns=["year", "t1d_registered", "t2d_registered", "source"])

    df = df_raw[[year_col, type_col, count_col]].copy()
    df.columns = ["year", "diabetes_type", "registered_count"]

    df["registered_count"] = df["registered_count"].str.replace(",", "", regex=False).str.strip()
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df["registered_count"] = pd.to_numeric(df["registered_count"], errors="coerce")
    df = df.dropna(subset=["year", "registered_count"])

    t1d = (df[df["diabetes_type"].str.contains("1형", na=False)]
           .groupby("year")["registered_count"].sum().reset_index()
           .rename(columns={"registered_count": "t1d_registered"}))
    t2d = (df[df["diabetes_type"].str.contains("2형", na=False)]
           .groupby("year")["registered_count"].sum().reset_index()
           .rename(columns={"registered_count": "t2d_registered"}))

    out = t1d.merge(t2d, on="year", how="outer").sort_values("year").reset_index(drop=True)
    out["year"] = out["year"].astype("Int64")
    out["t1d_registered"] = out["t1d_registered"].astype("Int64")
    out["t2d_registered"] = out["t2d_registered"].astype("Int64")
    out["source"] = "NHIS_YOYANGBI_REGISTERED_20241231"
    return out[["year", "t1d_registered", "t2d_registered", "source"]]


def parse_t1d_age_annual_csv(path: str | Path) -> pd.DataFrame:
    """Parse T1D patients by year × 1-year age band (2013–2023).

    Source: NHIS_T1D_AGE_ANNUAL_20231231. ~1,109 rows (11 years × ~101 age bands).
    Encoding: cp949.

    Suppression: cells with '*' represent patient counts < 5 — stored as NaN, not 0.
    '100세+' is mapped to age=100.

    Output: year (int), age (int), patients (float, nullable), suppressed (bool), source (str)
    """
    path = Path(path)

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read T1D age annual CSV: {e}")
        return pd.DataFrame(columns=["year", "age", "patients", "suppressed", "source"])

    df.columns = [c.strip() for c in df.columns]

    col_map = {}
    for c in df.columns:
        if "진료년도" in c or "진료연도" in c:
            col_map[c] = "year_raw"
        elif "연령" in c and "구분" not in c:
            col_map[c] = "age_raw"
        elif "진료인원" in c:
            col_map[c] = "patients_raw"
    df = df.rename(columns=col_map)

    if "year_raw" not in df.columns or "age_raw" not in df.columns or "patients_raw" not in df.columns:
        logger.error(f"Required columns missing from {path.name}. Columns: {list(df.columns)}")
        return pd.DataFrame(columns=["year", "age", "patients", "suppressed", "source"])

    # Parse year: strip '년' suffix
    df["year"] = df["year_raw"].str.replace("년", "", regex=False).str.strip()
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")

    # Parse age: strip '세+' / '세'; '100세+' → 100
    df["age_str"] = (df["age_raw"]
                     .str.replace("세+", "", regex=False)
                     .str.replace("세", "", regex=False)
                     .str.strip())
    df["age"] = pd.to_numeric(df["age_str"], errors="coerce").astype("Int64")

    # Suppression: '*' → NaN
    df["suppressed"] = df["patients_raw"].str.strip() == "*"
    df["patients"] = df["patients_raw"].where(~df["suppressed"], other=pd.NA)
    df["patients"] = df["patients"].str.replace(",", "", regex=False).str.strip()
    df["patients"] = pd.to_numeric(df["patients"], errors="coerce")

    df["source"] = "NHIS_T1D_AGE_ANNUAL_20231231"
    return df[["year", "age", "patients", "suppressed", "source"]].reset_index(drop=True)


def parse_annual_diabetes_clinical_xlsx(path: str | Path) -> pd.DataFrame:
    """Parse Sheet 1 of annual diabetes info XLSX for E10-E14 age-split data (2010–2023).

    Sheet: '당뇨병 연도별 연령별 진료인원'
    Source: NHIS_ANNUAL_DIABETES_CLINICAL_20241231. ~158 rows.

    MANDATORY CAVEAT: Pre-2018 E10 counts reflect ICD coding drift artifact —
    do not compare pre-2018 to post-2020 values directly.

    Output: year (int), icd_code (str), age_bracket (str), patient_count (int), source (str)
    """
    path = Path(path)

    df_raw = _read_nhis_xlsx_sheet(path, sheet_idx=0)
    if df_raw.empty:
        return pd.DataFrame(columns=["year", "icd_code", "age_bracket", "patient_count", "source"])

    year_col = next((c for c in df_raw.columns if "년도" in c or "연도" in c), None)
    icd_col = next((c for c in df_raw.columns if "상병코드" in c or c.upper() == "ICD"), None)
    age_col = next((c for c in df_raw.columns if "연령" in c and "구분" in c), None)
    count_col = next((c for c in df_raw.columns if "진료인원" in c), None)

    if not all([year_col, icd_col, age_col, count_col]):
        logger.warning(f"Column detection failed. Columns: {list(df_raw.columns)} — trying positional")
        cols = df_raw.columns.tolist()
        if len(cols) >= 4:
            year_col, icd_col, age_col, count_col = cols[0], cols[1], cols[2], cols[3]
        else:
            return pd.DataFrame(columns=["year", "icd_code", "age_bracket", "patient_count", "source"])

    df = df_raw[[year_col, icd_col, age_col, count_col]].copy()
    df.columns = ["year", "icd_code", "age_bracket", "patient_count"]

    df["patient_count"] = df["patient_count"].str.replace(",", "", regex=False).str.strip()
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df["patient_count"] = pd.to_numeric(df["patient_count"], errors="coerce")
    df = df.dropna(subset=["year", "patient_count"])
    df["year"] = df["year"].astype("Int64")
    df["patient_count"] = df["patient_count"].astype("Int64")
    df["source"] = "NHIS_ANNUAL_DIABETES_CLINICAL_20241231"
    return df[["year", "icd_code", "age_bracket", "patient_count", "source"]].reset_index(drop=True)


def parse_sigungu_t1d_t2d_xlsx(path: str | Path) -> pd.DataFrame:
    """Parse 시군구-level T1D+T2D XLSX (sex sheet and age sheet combined).

    Reads BOTH sheets from XLSX #15145378. Years: 2022–2024. ~247 시군구.
    Source: NHIS_SIGUNGU_T1D_T2D_20241231.

    Sheet handling:
      Sheet 0 (성별): dimension = sex value ('남', '여', '계')
      Sheet 1 (연령별): dimension = age value; '0세' normalized to '10대미만'

    Suppression: '*' → NaN. patients is NaN where suppressed=True.

    Output: sheet (str), 구분 (str), year (int), 시도 (str), 시군구 (str),
            dimension (str), patients (float, nullable), suppressed (bool), source (str)
    """
    path = Path(path)

    all_frames = []

    for sheet_idx, sheet_label in [(0, "sex"), (1, "age")]:
        df_raw = _read_nhis_xlsx_sheet(path, sheet_idx=sheet_idx)
        if df_raw.empty:
            logger.warning(f"Sheet {sheet_idx} empty or unreadable in {path.name}")
            continue

        # Find required columns
        type_col = next((c for c in df_raw.columns if c in ("구분",) or "구분" == c), None)
        year_col = next((c for c in df_raw.columns if "년도" in c or "연도" in c), None)
        sido_col = next((c for c in df_raw.columns if c == "시도" or "시도" in c and "시군구" not in c), None)
        sigungu_col = next((c for c in df_raw.columns if "시군구" in c), None)
        # Dimension column: sex or age
        dim_col = next((c for c in df_raw.columns if "성별" in c or "연령" in c), None)
        count_col = next((c for c in df_raw.columns if "환자수" in c or "진료인원" in c), None)

        if not all([year_col, sido_col, sigungu_col, dim_col, count_col]):
            logger.warning(f"Sheet {sheet_idx}: column detection failed. Columns: {list(df_raw.columns)}")
            continue

        df = df_raw[[
            *([type_col] if type_col else []),
            year_col, sido_col, sigungu_col, dim_col, count_col,
        ]].copy()

        if type_col:
            df.columns = ["구분", "year", "시도", "시군구", "dimension", "patients_raw"]
        else:
            df.columns = ["year", "시도", "시군구", "dimension", "patients_raw"]
            df.insert(0, "구분", "")

        df["suppressed"] = df["patients_raw"].str.strip() == "*"
        df["patients"] = df["patients_raw"].where(~df["suppressed"], other=pd.NA)
        df["patients"] = df["patients"].str.replace(",", "", regex=False).str.strip()
        df["patients"] = pd.to_numeric(df["patients"], errors="coerce")

        df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")

        # Normalize 0세 artifact in age sheet
        if sheet_label == "age":
            df["dimension"] = df["dimension"].str.replace("0세", "10대미만", regex=False)

        df["sheet"] = sheet_label
        df["source"] = "NHIS_SIGUNGU_T1D_T2D_20241231"
        df = df.drop(columns=["patients_raw"])

        all_frames.append(df[["sheet", "구분", "year", "시도", "시군구", "dimension",
                               "patients", "suppressed", "source"]])

    if not all_frames:
        return pd.DataFrame(columns=["sheet", "구분", "year", "시도", "시군구",
                                     "dimension", "patients", "suppressed", "source"])
    return pd.concat(all_frames, ignore_index=True)


def parse_t2d_sigungu_csv(path: str | Path) -> pd.DataFrame:
    """Parse T2D clinical data by institution type per 시군구 (2021–2023).

    Applies SIDO_NAME_MAP to normalize abbreviated 시도 names to full names.
    Source: NHIS_T2D_SIGUNGU_CLINICAL_20231231.

    Output: year (int), coverage_type (str), sido (str), sigungu (str),
            institution_type (str), patient_count (int), visit_count (int),
            cost_krw_thousands (float), source (str)
    """
    path = Path(path)

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read T2D 시군구 CSV: {e}")
        return pd.DataFrame(columns=["year", "coverage_type", "sido", "sigungu",
                                     "institution_type", "patient_count",
                                     "visit_count", "cost_krw_thousands", "source"])

    df.columns = [c.strip() for c in df.columns]

    col_map = {}
    for c in df.columns:
        if c == "구분":
            col_map[c] = "coverage_type"
        elif "년도" in c or "연도" in c:
            col_map[c] = "year"
        elif "시도" in c and "시군구" not in c:
            col_map[c] = "sido"
        elif "시군구" in c:
            col_map[c] = "sigungu"
        elif "기관종별" in c or "기관종류" in c or "종별" in c:
            col_map[c] = "institution_type"
        elif "진료인원" in c:
            col_map[c] = "patient_count"
        elif "진료건수" in c:
            col_map[c] = "visit_count"
        elif "진료비" in c:
            col_map[c] = "cost_krw_thousands"
    df = df.rename(columns=col_map)

    # Drop unneeded columns (주상병코드, 주상병명)
    keep = [c for c in ["coverage_type", "year", "sido", "sigungu", "institution_type",
                         "patient_count", "visit_count", "cost_krw_thousands"]
            if c in df.columns]
    df = df[keep].copy()

    # Normalize 시도 names
    if "sido" in df.columns:
        df["sido"] = df["sido"].replace(SIDO_NAME_MAP)

    # Parse numerics
    for col in ["patient_count", "visit_count", "cost_krw_thousands"]:
        if col in df.columns:
            df[col] = df[col].str.replace(",", "", regex=False).str.strip()
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "year" in df.columns:
        df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")

    df["source"] = "NHIS_T2D_SIGUNGU_CLINICAL_20231231"

    out_cols = ["year", "coverage_type", "sido", "sigungu", "institution_type",
                "patient_count", "visit_count", "cost_krw_thousands", "source"]
    available = [c for c in out_cols if c in df.columns]
    return df[available].reset_index(drop=True)


def parse_diabetes_utilization_rate_csvs(paths: list[str | Path]) -> pd.DataFrame:
    """Merge all 당뇨병의료이용률 CSVs into a single deduplicated DataFrame (2002–2024).

    All 6 source files share an identical 7-column schema (cp949 encoding):
      지표연도 → year, 시도 → sido, 시군구 → sigungu, 지표명 → indicator_name,
      분모(명) → denominator, 분자(명) → numerator, 지표값(퍼센트) → utilization_rate_pct

    Merge strategy: read all files, concat, deduplicate on (year, sido, sigungu) keeping
    the last occurrence (sorted filenames → newer files take precedence at edge years).

    Output schema:
        year (int), sido (str), sigungu (str), indicator_name (str),
        denominator (int), numerator (int), utilization_rate_pct (float), source (str)
    """
    EMPTY_COLS = ["year", "sido", "sigungu", "indicator_name",
                  "denominator", "numerator", "utilization_rate_pct", "source"]

    col_map = {
        "지표연도": "year",
        "시도": "sido",
        "시군구": "sigungu",
        "지표명": "indicator_name",
        "분모(명)": "denominator",
        "분자(명)": "numerator",
        "지표값(퍼센트)": "utilization_rate_pct",
    }

    # Paths are pre-sorted by find_diabetes_utilization_csvs(); no re-sort needed.
    frames: list[pd.DataFrame] = []
    for raw_path in paths:
        path = Path(raw_path)
        try:
            df = pd.read_csv(path, encoding="cp949", dtype=str)
        except Exception as e:
            logger.error(f"Failed to read {path.name}: {e}")
            continue

        df.columns = [c.strip() for c in df.columns]
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

        # Keep only mapped columns that exist; warn if any expected column is absent.
        expected = [c for c in EMPTY_COLS if c != "source"]
        keep = [c for c in expected if c in df.columns]
        missing = set(expected) - set(keep)
        if missing:
            logger.warning(f"{path.name}: missing expected columns {sorted(missing)}")
        df = df[keep].copy()
        df["source"] = "NHIS_DIABETES_UTILIZATION_RATE"
        frames.append(df)

    if not frames:
        return pd.DataFrame(columns=EMPTY_COLS)

    combined = pd.concat(frames, ignore_index=True)

    # Parse numerics
    for col in ["year", "denominator", "numerator"]:
        if col in combined.columns:
            combined[col] = (combined[col]
                             .str.replace(",", "", regex=False).str.strip())
            combined[col] = pd.to_numeric(combined[col], errors="coerce").astype("Int64")

    if "utilization_rate_pct" in combined.columns:
        combined["utilization_rate_pct"] = (
            combined["utilization_rate_pct"]
            .str.replace(",", "", regex=False).str.strip()
        )
        combined["utilization_rate_pct"] = pd.to_numeric(
            combined["utilization_rate_pct"], errors="coerce"
        )

    # Deduplicate: keep last (newest file wins at overlapping years)
    dedup_keys = [c for c in ["year", "sido", "sigungu"] if c in combined.columns]
    combined = combined.drop_duplicates(subset=dedup_keys, keep="last")

    combined = combined.dropna(subset=["year"]).sort_values(
        ["year", "sido", "sigungu"], na_position="last"
    ).reset_index(drop=True)

    available = [c for c in EMPTY_COLS if c in combined.columns]
    return combined[available]


def parse_insulin_claims_csv(path: str | Path) -> pd.DataFrame:
    """Parse 당뇨병 진료인원 중 인슐린 주사 청구건수 및 금액 CSV.

    Monthly insulin injection claims by age group, 2016–2023. 865 rows.
    Encoding: cp949.

    Schema:
      진료년도 → year, 진료월 → month, 연령 → age_group,
      청구건수(건) → claim_count, 청구금액(천원) → claim_amount_krw_thousands

    NOTE: Covers ALL insulin-dependent diabetics (T1D + insulin-dependent T2D),
    not T1D alone. Cannot be attributed to CGM users specifically.

    Output schema:
        year (int), month (int), age_group (str), claim_count (int),
        claim_amount_krw_thousands (float), source (str)
    """
    EMPTY_COLS = ["year", "month", "age_group", "claim_count",
                  "claim_amount_krw_thousands", "source"]
    path = Path(path)

    try:
        df = pd.read_csv(path, encoding="cp949", dtype=str)
    except Exception as e:
        logger.error(f"Failed to read insulin claims CSV: {e}")
        return pd.DataFrame(columns=EMPTY_COLS)

    df.columns = [c.strip() for c in df.columns]

    # Column detection by substring
    year_col = next((c for c in df.columns if "진료년도" in c or "진료연도" in c), None)
    month_col = next((c for c in df.columns if "진료월" in c), None)
    age_col = next((c for c in df.columns if "연령" in c), None)
    count_col = next((c for c in df.columns if "청구건수" in c), None)
    amount_col = next((c for c in df.columns if "청구금액" in c), None)

    if not all([year_col, month_col, age_col, count_col, amount_col]):
        logger.error(
            f"Cannot find required columns in {path.name}. Columns: {list(df.columns)}"
        )
        return pd.DataFrame(columns=EMPTY_COLS)

    df = df.rename(columns={
        year_col: "year",
        month_col: "month",
        age_col: "age_group",
        count_col: "claim_count",
        amount_col: "claim_amount_krw_thousands",
    })

    for col in ["year", "month", "claim_count"]:
        df[col] = (df[col].str.replace(",", "", regex=False).str.strip())
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    df["claim_amount_krw_thousands"] = (
        df["claim_amount_krw_thousands"].str.replace(",", "", regex=False).str.strip()
    )
    df["claim_amount_krw_thousands"] = pd.to_numeric(
        df["claim_amount_krw_thousands"], errors="coerce"
    )

    df["source"] = "NHIS_INSULIN_CLAIMS_20231231"
    df = df.dropna(subset=["year", "month"]).copy()

    return df[EMPTY_COLS].reset_index(drop=True)



def download_publication_stats(year: int, output_dir: Path | None = None) -> pd.DataFrame:
    """NOT AUTOMATABLE — NHIS publication stats require manual download.

    NHIS 발간자료 (dataset #15095102) are Excel files that must be downloaded
    manually from nhiss.nhis.or.kr — they are NOT accessible via REST API.
    Use parse_yearbook_ch06() for programmatic access to equivalent data.
    """
    raise NotImplementedError(
        "NHIS publication stats cannot be fetched automatically. "
        "Download Excel files manually from nhiss.nhis.or.kr "
        "(발간자료 → 건강보험통계연보) and parse with parse_yearbook_ch06()."
    )


def _extract_year_from_path(path: Path) -> int | None:
    """Extract 4-digit year from a file path — searches all path components.

    Searches stem, name, and ALL parent directory names to handle deep paths like:
      (본문 및 해설서)2024 건강보험통계연보/1. 본문/ch06.xlsx
    where the year is two directory levels above the filename.
    """
    for part in [path.stem, path.name] + [p.name for p in path.parents]:
        m = re.search(r"(20[12]\d)", part)
        if m:
            return int(m.group(1))
    return None
